import openpyxl
import time
import string

workbook = openpyxl.load_workbook(r'C:\Users\Daniel\Desktop\autoformBot\DummyData.xlsx')
worksheet = workbook['Sheet3']

# cell_alpabet = 'A'
# cell_num=2

#CHECKING IF CELL IS EMPTY AND COUNT IT WITH CELL_NUM
# while True:
#     cell_value = worksheet['A' + str(cell_num)].value
#     if cell_value is not None:
#         cell_num +=1
#         time.sleep(1)
#         print(cell_value)
#     else:
#         break
# print(cell_num) 

#column -> horizontal
#row -> vertical

cell_col = worksheet.max_column
cell_row = worksheet.max_row

print(f'column: {cell_col}') 
print(f'row: {cell_row}')
#Reading cell 

# for i in range(2,cell_row):
#     for letter in  string.ascii_uppercase[:11]:
#         print(worksheet[str(letter) + str(i)].value)



import string

# Define a dictionary mapping column letters to variable names
variable_mapping = {
    'A': 'name',
    'B': 'block_number',
    'C': 'floor_number',
    'D': 'unit_number',
    'E': 'postal_code',
    'F': 'training',
    'G': 'status',
    'H': 'date_string',
    'I': 'officer_name',
    'J': 'NPC'
}

# Initialize variables
name = block_number = floor_number = unit_number = postal_code = training = status = date_string = officer_name = NPC = None

# Iterate over rows starting from row 2
for i in range(2, cell_row):
    for letter, var_name in variable_mapping.items():
        # Get the value from the corresponding cell
        cell_value = worksheet[letter + str(i)].value
        # Assign the value to the corresponding variable
        globals()[var_name] = cell_value

    # Do something with the variables, for example, print them
    print(name, block_number, floor_number, unit_number, postal_code, training, status, date_string, officer_name, NPC)





# Iterate through columns



# # Close the workbook
# workbook.close()

