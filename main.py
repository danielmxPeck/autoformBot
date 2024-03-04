from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import string
import openpyxl
import time
import string

options = webdriver.ChromeOptions()
options.add_experimental_option("detach",True)
url = "https://form.gov.sg/657903cb2086120011606261"
driver = webdriver.Chrome(options=options, service=Service(ChromeDriverManager().install()))
driver.implicitly_wait(10)
driver.get(url)

responseID = []

# def input(name, block_number, floor_number, unit_number, postal_code, training, status, date ,officer_name, NPC):
def input(name , block_number, floor_number, unit_number, postal_code, training, status, date, officer_name, NPC):
    driver.find_element(By.ID, '5e960d0a78a996001147755b').send_keys(name)
    driver.find_element(By.ID, '646741312034d2001235bdac').send_keys(block_number)
    driver.find_element(By.ID, '646741581e17d000127b9f3b').send_keys(floor_number)
    driver.find_element(By.ID, '646741712014ed0011cdd4a4').send_keys(unit_number)
    driver.find_element(By.ID, '64674227c36fce00121019ba').send_keys(postal_code)
    driver.find_element(By.ID, '64674334dcd7850012786996').send_keys(training, Keys.TAB)
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, '659ce77a96c159001278539c'))).send_keys(status, Keys.TAB)
    date_input = driver.find_element(By.ID, '646743551e17d000127beb5b')
    date_input.clear()
    date_input.send_keys(Keys.BACKSPACE * 10)
    date_input.send_keys(date)
    driver.find_element(By.ID, '6467436530fc910012d93ed3').send_keys(officer_name)
    driver.find_element(By.ID, '6467453530fc910012d97feb').send_keys(NPC, Keys.TAB)
    driver.find_element(By.CLASS_NAME, 'chakra-button.css-1szjd8b').click()  

    response_id_element = driver.find_element(By.CLASS_NAME, 'chakra-text.css-jdumc1')
    response_id_text = response_id_element.text
    response_id = response_id_text.split(": ")[1]
    responseID.append(response_id)
    # Click on the button
    print(responseID)
    another_form = driver.find_element(By.CLASS_NAME, 'chakra-button.css-4my61l')
    another_form.click()


workbook = openpyxl.load_workbook(r'C:\Users\Daniel\Desktop\autoformBot\DummyData.xlsx')
worksheet = workbook['Sheet3']

cell_col = worksheet.max_column
cell_row = worksheet.max_row
 
# Define a dictionary mapping column letters to variable names
variable_mapping = {
    'A': 'name',
    'B': 'block_number',
    'C': 'floor_number',
    'D': 'unit_number',
    'E': 'postal_code',
    'F': 'training',
    'G': 'status',
    'H': 'date_string',
    'I': 'officer_name',
    'J': 'NPC'
}

# Initialize variables
name = block_number = floor_number = unit_number = postal_code = training = status = date_string = officer_name = NPC = None

# Iterate over rows starting from row 2
for i in range(2, cell_row):
    for letter, var_name in variable_mapping.items():
        # Get the value from the corresponding cell
        cell_value = worksheet[letter + str(i)].value
        # Assign the value to the corresponding variable
        globals()[var_name] = cell_value

    # Do something with the variables, for example, print them
    input(name, block_number, floor_number, unit_number, postal_code, training, status, date_string, officer_name, NPC)


# Close the workbook
workbook.close()
 